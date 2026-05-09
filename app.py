from PIL import Image
import streamlit as st
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
import openpyxl
import os
from datetime import datetime
from textwrap import wrap
import pytz
import gspread
import json
from oauth2client.service_account import ServiceAccountCredentials
from io import BytesIO

st.set_page_config(page_title="Sales Order", layout="centered")
st.title("📦 Sales Order Generator")

def get_order_number(prefix):
    file = "order_series.xlsx"

    if not os.path.exists(file):
        wb = openpyxl.Workbook()
        sheet = wb.active

        sheet.append(["Type","Last Number"])
        sheet.append(["SO",14500])
        sheet.append(["SM",14000])
        sheet.append(["SOT",50000])
        sheet.append(["VFSO",10000])
        sheet.append(["VFSM",10000])

        wb.save(file)

    wb = openpyxl.load_workbook(file)
    sheet = wb.active

    for row in sheet.iter_rows(min_row=2):

        if row[0].value == prefix:
            row[1].value += 1

            num = row[1].value

            wb.save(file)

            return f"{prefix} {num}"

# ================= FORM =================

name = st.text_input("Party Name")
address = st.text_input("Address")
salesman = st.text_input("Salesman")

order_type = st.selectbox(
    "Order Type",
    ["SO","SM","SOT","VFSO","VFSM"]
)

manual_no = st.text_input("Manual Order No")

old_order_check = st.checkbox("Old Order")

old_order_no = st.text_input(
    "Enter Old Order No"
) if old_order_check else ""

# ================= FABRIC =================

st.subheader("Fabric Setup")

fabric1 = st.text_input("Fabric 1")
rate1 = st.number_input("Rate 1", min_value=0)

fabric2 = st.text_input("Fabric 2")
rate2 = st.number_input("Rate 2", min_value=0)

fabric3 = st.text_input("Fabric 3")
rate3 = st.number_input("Rate 3", min_value=0)

fabric_options = [
    f for f in [fabric1, fabric2, fabric3] if f
]

# ================= IMAGE =================

uploaded_files = st.file_uploader(
    "Upload Designs",
    type=["jpg","jpeg","png"],
    accept_multiple_files=True
)

design_data = []

if uploaded_files:

    for i, file in enumerate(uploaded_files):

        st.markdown(f"### Design {i+1}")

        st.image(file, width=150)

        unit = st.selectbox(
            f"Unit {i+1}",
            ["PCS","MTR"],
            key=f"unit{i}"
        )

        qty = st.number_input(
            f"{unit} {i+1}",
            min_value=1.0,
            key=f"qty{i}"
        )

        fabric = st.selectbox(
            f"Fabric {i+1}",
            fabric_options,
            key=f"fab{i}"
        )

        measurement_type = st.selectbox(
           f"Type {i+1}",
           ["Cut", "Length"],
           key=f"type{i}"
        )

        cut = st.number_input(
           f"{measurement_type} {i+1}",
           min_value=0.0,
           key=f"cut{i}"
        )

        design_description = st.text_area(
           f"Description {i+1}",
           key=f"desc{i}"
        )

        mtr = qty * cut if unit == "PCS" else qty

        design_data.append({
    "file": file,
    "pcs": qty if unit == "PCS" else 0,
    "mtr": mtr,
    "unit": unit,
    "fabric": fabric,
    "cut": cut,
    "type": measurement_type,
    "description": design_description
})

description = st.text_area("Description")

ref_image = st.file_uploader(
    "Upload Reference Image (Optional)",
    type=["jpg","jpeg","png"]
)

# ================= PDF =================

def create_pdf(data, design_data):

    filename = f"{data['order_no']}.pdf"

    c = canvas.Canvas(filename, pagesize=A4)

    width, height = A4

    def header():

        y = height - 40

        c.setFont("Helvetica-Bold", 14)

        label_x = width / 2 - 120
        value_x = width / 2 - 10

        c.drawString(label_x, y, "Date:")
        c.drawString(value_x, y, str(data["date"]))

        c.drawString(label_x, y - 20, "ORDER NO:")
        c.drawString(value_x, y - 20, str(data["order_no"]))

        c.drawString(label_x, y - 40, "Salesman:")
        c.drawString(value_x, y - 40, str(data["salesman"]))

        if data["old_order"]:
            c.drawString(label_x, y - 60, "Old Order:")
            c.drawString(value_x, y - 60, str(data["old_order"]))

        return y-90

    y = header()

    current_y = y

    row_max_height = 0

    col_index = 0

    for i, d in enumerate(design_data):

        x = 60 if col_index % 2 == 0 else 300

        base_y = current_y

        img = Image.open(d["file"])

        iw, ih = img.size

        ratio = min(180/iw, 180/ih)

        new_w = iw * ratio
        new_h = ih * ratio

        required_height = new_h + 80

        if (
            col_index % 2 == 0 and
            base_y - required_height < 40
        ):

            c.showPage()

            y = header()

            current_y = y

            row_max_height = 0

            col_index = 0

            x = 60

            base_y = current_y

        c.setFont("Helvetica-Bold", 12)
        c.drawString(x, base_y + 15, f"{i+1}.")

        buffer = BytesIO()

        rgb_img = img.convert("RGB")
        rgb_img.save(
            buffer,
            format="JPEG",
            quality=88,
            optimize=True
        )

        buffer.seek(0)

    c.drawImage(
        ImageReader(img),
        x,
        base_y - new_h,
        width=new_w,
        height=new_h
    )            
    text_y = base_y - new_h - 10
    
    c.setFont("Helvetica-Bold", 12)
    c.drawString(x, text_y, d["fabric"])

    if d["unit"] == "PCS":

            c.drawString(
                x,
                text_y-15,
                f"PCS: {d['pcs']}"
            )

            c.drawString(
                x,
                text_y-30,
                f"{d['type']}: {d['cut']}"
            )

            extra_desc_height = 0

            if d["description"]:

                desc_lines = wrap(d["description"], 28)

                desc_y = text_y - 45

                for line in desc_lines:

                    c.setFont("Helvetica", 10)

                    c.drawString(
                        x,
                        desc_y,
                        line
                    )

                    desc_y -= 12

                extra_desc_height = len(desc_lines) * 12

    else:

            c.drawString(
                x,
                text_y-15,
                f"MTR: {d['mtr']}"
            )

            c.setFont("Helvetica", 10)
            
            extra_desc_height = 0

            if d["description"]:

                desc_lines = wrap(d["description"], 28)

                desc_y = text_y - 30
                
                c.setFont("Helvetica", 12)

                for line in desc_lines:

                    c.drawString(
                        x,
                        desc_y,
                        line
                    )

                    desc_y -= 12

            c.setFont("Helvetica", 12)        
                    

    extra_desc_height = len(desc_lines) * 12

    total_height = new_h + 60 + extra_desc_height

    if total_height > row_max_height:
            row_max_height = total_height

    col_index += 1

    if col_index % 2 == 0:

            current_y -= (row_max_height + 30)

            row_max_height = 0

    # ================= DESCRIPTION =================

    if description:

        c.showPage()

        y = 750

        # IMAGE FIXED

        if ref_image:

            image = Image.open(ref_image)

            iw, ih = image.size

            max_w, max_h = 300, 300

            ratio = min(max_w/iw, max_h/ih)

            new_w = iw * ratio
            new_h = ih * ratio

            x = (width - new_w) / 2

            c.drawImage(
                ImageReader(image),
                x,
                y - new_h,
                width=new_w,
                height=new_h
            )

            y = y - new_h - 30

        # DESCRIPTION TEXT

        lines = wrap(description, 80)

        c.setFillColorRGB(1,0,0)

        c.setFont("Helvetica-Bold", 10)

        for line in lines:

            c.drawCentredString(width/2, y, line)

            y -= 15

    c.save()

    return filename

# ================= GOOGLE SHEET =================

def save_to_google_sheet(data, design_data):

    scope = [
        "https://spreadsheets.google.com/feeds",
        "https://www.googleapis.com/auth/drive"
    ]

    creds_dict = json.loads(st.secrets["GOOGLE_CREDS"])

    creds = ServiceAccountCredentials.from_json_keyfile_dict(
        creds_dict,
        scope
    )

    client = gspread.authorize(creds)

    sheet = client.open_by_key(
        "1FTshf42DtsPfVav6RTlHzK2g_AOZF2ZUxBl7swTe2Xk"
    ).sheet1

    fabrics = ["", "", ""]
    mtrs = [0, 0, 0]

    for d in design_data:

        if d["fabric"] in fabrics:

            idx = fabrics.index(d["fabric"])

            mtrs[idx] += d["mtr"]

        else:

            for i in range(3):

                if fabrics[i] == "":

                    fabrics[i] = d["fabric"]

                    mtrs[i] = d["mtr"]

                    break

    total_mtr = sum(mtrs)

    row = [
        data["order_no"],
        data["name"],
        data["address"],
        data["date"],
        data["salesman"],

        fabrics[0], mtrs[0], "",
        fabrics[1], mtrs[1], "",
        fabrics[2], mtrs[2], "",

        total_mtr
    ]

    sheet.append_row(row)

# ================= BUTTON =================

if st.button("🚀 Generate Order"):

    order_no = (
        f"{order_type} {manual_no}"
        if manual_no
        else get_order_number(order_type)
    )

    india = pytz.timezone("Asia/Kolkata")

    current_time = datetime.now(india)

    data = {
        "order_no": order_no,
        "name": name,
        "address": address,
        "salesman": salesman,
        "date": current_time.strftime("%d/%m/%y"),
        "old_order": old_order_no
    }

    pdf = create_pdf(data, design_data)

    # SAVE TO GOOGLE SHEET
    save_to_google_sheet(data, design_data)

    st.success("Order Created")

    with open(pdf, "rb") as f:

        st.download_button(
            "Download PDF",
            f,
            file_name=pdf
        )
